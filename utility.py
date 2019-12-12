from openpyxl import load_workbook, Workbook

def load_data(file):
    """Load dataset as two dimesional list. Columns save 
    every feature's observation value, and rows save a 
    record of all features. Last column is class, or label, 
    of answer.

    Args:
        file: path of an xlsx format file
    Return:
        dataSet: Same as above description
        labels: Name of feature 

    """
    wb = load_workbook(file)
    ws = wb.active

    dataSet = []
    labels = None

    for i,row in enumerate(ws.iter_rows()):
        rowdata = [cell.value for cell in row]
        if i > 1:
            rowdata = rowdata[1:]
            c = rowdata[0]
            del rowdata[0]
            rowdata.append(c)
            dataSet.append(rowdata)
        elif i == 1:
            rowdata = rowdata[2:]
            for i,r in enumerate(rowdata):
                rowdata[i] = rowdata[i][1:-1]
            labels = rowdata
    
    return dataSet, labels