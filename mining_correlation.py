import argparse
import math
from pathlib import Path
import os

from openpyxl import load_workbook, Workbook
import numpy as np


parser = argparse.ArgumentParser()
parser.add_argument('--xlsx', help='need a xlsx file. \
                    e.g. --xlsx example.xlsx', dest='XLSX')
args = parser.parse_args()

def pair_Pearson_Correlation(X,Y):
    """
    X and Y are two attributes, len(X) == len(Y)
    PPC is to calculate these two attributes' correlation
    if output value:
        1) is more closer to positive one(+1): 
             X, Y are positive highly correlate
        2) is more closer to zero:
             X, Y are not very correlate
        3) is more closer to negative one(-1):
             X, Y are negative highly correlate
    PPC(X, Y) = Cov(X, Y) / (SD(X)*SD(Y))

    Cov: Covariance
    SD: Standard Deviation
    """
    cov = sum((X-np.mean(X))*(Y-np.mean(Y)))
    dex = math.sqrt(sum(np.square(X - np.mean(X))))
    dey = math.sqrt(sum(np.square(Y - np.mean(Y))))
    denom = dex * dey
    ppc = cov / denom
    return np.round(ppc, 2)

wb = load_workbook(args.XLSX)
ws = wb.active
#Subtypes = np.array([row[1].value for i, row
#                 in enumerate(ws.iter_rows()) if i > 1])


Features = np.zeros(110)
F_Name = []
for k in range(2,22):
    f = np.array([row[k].value for i, row 
        in enumerate(ws.iter_rows()) if i > 1], dtype='float32')
    name = [row[k].value for i, row 
        in enumerate(ws.iter_rows()) if i == 1]

    Features = np.vstack((Features, f))
    F_Name.append(name[0][1:-1])
Features = np.delete(Features, np.s_[0:1], axis=0)


wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'correlation of 20 Features'


offset = 2
for row in range(1, 22):
    for col in range(1, 22):
        if row == 1 and col == 1:
            pass
        elif row == 1 and col < 22:
            ws.cell(row=row, column=col).value = F_Name[col-offset]
        elif col == 1 and col < 22:
            ws.cell(row=row, column=col).value = F_Name[row-offset]
        else:
            ppc = pair_Pearson_Correlation(Features[row-offset], Features[col-offset])
            ws.cell(row=row, column=col).value = ppc
 



savepath = args.XLSX.split('/')
savepath = Path(os.path.join(*savepath[:-1]))
print('save correlation at:', savepath)
wb.save(filename=savepath / 'pairwise_Pearson_Correlation.xlsx')




