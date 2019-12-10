import argparse
import pickle
from math import log

from openpyxl import load_workbook, Workbook

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn import metrics

parser = argparse.ArgumentParser()
parser.add_argument('--xlsx', help='need a xlsx file. \
                    e.g. --xlsx example.xlsx', dest='XLSX')
args = parser.parse_args()



def loadData(file):
    wb = load_workbook(file)
    ws = wb.active

    dataSet = []
    labels = None

    for i,row in enumerate(ws.iter_rows()):
        rowdata = [cell.value for cell in row]
        if i > 1:
            rowdata = rowdata[1:]
            c = rowdata[0]
            del rowdata[0]
            rowdata.append(c)
            dataSet.append(rowdata)
        elif i == 1:
            rowdata = rowdata[2:]
            for i,r in enumerate(rowdata):
                rowdata[i] = rowdata[i][1:-1]
            labels = rowdata
    
    return dataSet, labels

def impurity(groups, classes, mode='gini'):
    n_instance = float(sum([len(group) for group in groups]))

    gini = 0
    parent = [0] * len(classes)
    entropy = 0
    gainRatio = 0
    splitInfo = 0
    
    for group in groups:
        size = float(len(group))
        if size == 0: continue
        score = 0
        score_e = 0
        for class_val in classes:
            p = [row[-1] for row in group].count(class_val) / size
            
            if mode == 'gini':
                score += p*p
            elif mode == 'entropy':
                if p != 0:
                    score_e -= p*log(p,2)
            elif mode == 'gain_ratio':
                parent[class_val] += [row[-1] for row in group].count(class_val)

        if mode == 'gini':
            gini += (size/n_instance) * (1 - score)
        elif mode == 'entropy':
            entropy += (size/n_instance) * score_e
        elif mode == 'gain_ratio':
            p = size/n_instance
            if p != 0:
                splitInfo -= p*log(p,2)
    if mode == 'gini':
        return gini
    if mode == 'entropy':
        return entropy
    if mode == 'gain_ratio':
        entropy_p = 0
        for pt in parent:
            p = pt/n_instance
            if p != 0:
                entropy_p -= p*log(p,2)
        info_gain = entropy_p - entropy
    
        if splitInfo != 0:
            gainRatio = info_gain/splitInfo
        
        return gainRatio


def continuous_attribute_split_position(featList):
    """
    input 7 values, output 8 values
     1 2 3 4 5 6 7
    ^ ^ ^ ^ ^ ^ ^ ^
    """
    featList.sort()

    s = []
    s.append(featList[0] - (featList[0]+featList[1])/2)
    for i in range(0, len(featList)-1):
        s.append((featList[i]+featList[i+1])/2)
    s.append(featList[-1] + (featList[0]+featList[1])/2)
    return s
#------------------

def test_split(dataset, index, value):     
    left = []
    right = []                                     
    for row in dataSet:
        if row[index] < value:
            left.append(row)
        elif row[index] >= value:
            right.append(row)
    return left, right

def Dataset_class(dataSet):
    totalNum = len(dataSet)
    labelNum = {}
    for data in dataSet:
        class_ = data[1]
        if class_ in labelNum:
            labelNum[class_] += 1
        else:
            labelNum[class_] = 1

def get_best_feature_to_split(dataSet):
    numFeatures = len(dataSet[0]) - 2                  #特征数量
    parentImpurity = cal_2class_impurity(dataSet)
    print('feature number: ',numFeatures) 
    print('parentImpurtiy: ', parentImpurity)
    

    bestInfoGain = 0.0                                #信息增益
    bestFeature = -1
    bestSplitValue = 0.0                                    #最优特征的索引值
    for i in range(2, numFeatures):                        #遍历所有特征
        #获取dataSet的第i个所有特征
        featList = [example[i] for example in dataSet]
        splitPositions = continuousAttribute(featList)                    
        newImpurity = 100000.0
        newSplitValue = 0.0                 
        for value in splitPositions:     
            subL, subR = splitDataSet(dataSet, i, value)      #subDataSet划分后的子集

            probL = len(subL) / float(len(dataSet))        #计算子集的概率
            probR = len(subR) / float(len(dataSet))
            giniSplit = probL * cal_2class_impurity(subL) \
                      + probR * cal_2class_impurity(subR)

            #print(value, ' : ', len(subL), ' ', len(subR), ', giniSplit: ', giniSplit)
            if giniSplit < newImpurity:
                newImpurity = giniSplit
                newSplitValue = value
        

        infoGain = parentImpurity - newImpurity                     #信息增益
        #print("第%d个特征的增益为%.3f, gini split值為%.3f, 最低gini的split position為%f" % (i-2, infoGain, newEntropy, newSplitValue))           #打印每个特征的信息增益
        if (infoGain > bestInfoGain):                           #计算信息增益
            bestInfoGain = infoGain                             #更新信息增益，找到最大的信息增益
            bestFeature = i-2
            bestSplitValue = newSplitValue


    #print("最終：第%d个特征的增益为%.3f, 最低gini的split position為%f" % (bestFeature, bestInfoGain, bestSplitValue))  
    return bestFeature, bestSplitValue





def majorityCnt(classList):
    classCount = {}
    for vote in classList:  #统计classList中每个元素出现的次数
        if vote not in classCount.keys():classCount[vote] = 0   
        classCount[vote] += 1
    sortedClassCount = sorted(classCount.items(), key = operator.itemgetter(1), reverse = True)     #根据字典的值降序排序
    return sortedClassCount[0][0]   #返回classList中出现次数最多的元素

def createTree(dataSet, labels, featLabels):
    classList = None
    try:
        classList = [example[1] for example in dataSet]
        if classList.count(classList[0]) == len(classList):         #如果类别完全相同则停止继续划分
            return classList[0]
        if len(dataSet[0]) == 1:                                    #遍历完所有特征时返回出现次数最多的类标签
            return majorityCnt(classList)
    except:
        print('classfication done.')
        return
    
    bestFeat, bestSplitValue = chooseBestFeatureToSplit(dataSet)
     
    print('best:', bestFeat) 
    print(labels)                #选择最优特征
    bestFeatLabel = labels[bestFeat]
    
    #print(bestSplitValue)                           #最优特征的标签
    featLabels.append(bestFeatLabel)
    bestFeatLabel +=  ' ' + str(bestSplitValue)
    myTree = {bestFeatLabel:{}}                                 #根据最优特征的标签生成树
    del(labels[bestFeat])
    print(myTree)
                                          
    #featValues = [example[bestFeat] for example in dataSet]     #得到训练集中所有最优特征的属性值
    #uniqueVals = set(featValues)                                #去掉重复的属性值
    subL, subR = splitDataSet(dataSet, bestFeat+2, bestSplitValue)
    left = '<= ' + str(bestSplitValue)
    right = '> ' + str(bestSplitValue)
    #print(len(subL))
    labelNum = cal_2class_impurity(subL, 'classCount')
    print(left, 'left',labelNum)
    #print(len(subR))
    labelNum = cal_2class_impurity(subR, 'classCount')
    print(right, 'right',labelNum)
    
    #print(len(dataSet))
    #print(len(subL))
    #print(len(subR))
    print('\n')
    if len(labels) > 0:                   
        myTree[bestFeatLabel]['left'] = createTree(subL, labels, featLabels)
        
        
    if len(labels) > 0:
        myTree[bestFeatLabel]['right'] = createTree(subR, labels, featLabels)
        
    
    
    return myTree

#------------------



#dataSet, labels, F_Name = loadData2(args.XLSX)
dataSet, labels,  = loadData(args.XLSX)

dataset = [[2.771244718,1.784783929,0],
    [1.728571309,1.169761413,0],
    [3.678319846,2.81281357,0],
    [3.961043357,2.61995032,0],
    [2.999208922,2.209014212,0],
    [7.497545867,3.162953546,1],
    [9.00220326,3.339047188,1],
    [7.444542326,0.476683375,1],
    [10.12493903,3.234550982,1],
    [6.642287351,3.319983761,1]]

#BASE = None
#if mode == 'gini' or 'entropy'
#BASE = 10000
#if mode == 'gain_ratio'
#BASE = 0
#classes = [0, 1]

#value = impurity(dataset, [0,1], mode='gain_ratio')
dataset = dataSet

class_values = list(set(row[-1] for row in dataset))

b_f_index, b_value, b_score, groups = None, None, 10000, None
for f_index in range(len(dataset[0])-1):
    
    column_value = [example[f_index] for example in dataset]
    split_values = continuous_attribute_split_position(column_value)

    
    for value in split_values:
        groups = test_split(dataset, f_index, value)
        score = impurity(groups, class_values, 'gini')
        #if mode == 'gini' or 'entropy':
        if score < b_score:
            b_f_index, b_value, b_score, b_groups = f_index, value, score, groups
        #if mode == 'gain_ratio':
        #if split_impurity > b_split_impurity:
        #    b_split_impurity = split_impurity
        print('feature %d < %f score = %.3f / best: feature %d < %f score = %.3f' % \
            (f_index, value, score, b_f_index, b_value, b_score))


#featLabels = []
#myTree = createTree(dataSet, labels, featLabels)

#print(myTree)


dataset = [[[2.771244718,1.784783929,0],
    [1.728571309,1.169761413,0],
    [3.678319846,2.81281357,0],
    [3.961043357,2.61995032,0],
    [2.999208922,2.209014212,0],
    [7.497545867,3.162953546,1],
    [9.00220326,3.339047188,1],
    [7.444542326,0.476683375,1],
    [10.12493903,3.234550982,1],
    [6.642287351,3.319983761,1]]]





