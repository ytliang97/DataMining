
from collections import deque

from openpyxl import load_workbook, Workbook
import numpy as np
from matplotlib import pyplot
from sklearn import tree

def calcGini(dataSet):
    totalNum = shape(dataSet)[0]
    labelNum = {}
    gini = 0
    for data in dataSet:
        label = data[-1]
        if label in labelNum:
            labelNum[label] += 1
        else:
            labelNum[label] = 1
 
    for key in labelNum:
        p = labelNum[key] / totalNum
        gini += p * (1 - p)
    return gini


def loadData(file):
    wb = load_workbook(file)
    ws = wb.active

    Data = np.zeros(110)
    F_Name = []
    Label = []
    for k in range(1,22):
        if k == 1:
            Label = np.array([row[k].value for i, row 
                in enumerate(ws.iter_rows()) if i > 1], dtype='int32')
        if k > 1:
            f = np.array([row[k].value for i, row 
                in enumerate(ws.iter_rows()) if i > 1], dtype='float32')
            Data = np.vstack((Data, f))
            name = [row[k].value for i, row 
                in enumerate(ws.iter_rows()) if i == 1 and k > 1]
        
            F_Name.append(name[0][1:-1])
    Data = np.delete(Data, np.s_[0:1], axis=0)

    
    return Data.T, Label, F_Name


def cal_2class_impurity(dataSet, label, mode='gini'):

    if mode == 'gini':
        totalNum = dataSet.shape[0]
        labelNum = {}
        gini = 0
        for data in dataSet:
            #label = data[-1]
            if label in labelNum:
                labelNum[label] += 1
            else:
                labelNum[label] = 1
     
        for key in labelNum:
            p = labelNum[key] / totalNum
            gini += p * (1 - p)
        return gini
    elif mode == 'entropy':
        pass
    elif mode == 'gain_ratio':
        pass



Data, Label, F_Name = loadData(args.XLSX)
gini = cal_2class_impurity(Data, Label, 'gini')

#impurity = cal_2class_impurity(class_vector)
#print(impurity)