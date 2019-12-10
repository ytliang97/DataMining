import argparse
#import math

from openpyxl import load_workbook, Workbook
import numpy as np
from matplotlib import pyplot
from sklearn.preprocessing import StandardScaler

parser = argparse.ArgumentParser()
parser.add_argument('--xlsx', help='need a xlsx file. \
                    e.g. --xlsx example.xlsx', dest='XLSX')
args = parser.parse_args()

def calculate_eigen_of_cov_mat(Attributes):
    """
    Need: import numpy as np
    Args:
        Attributes is a 2 dimensional array, and assume it is M x N. It means 
    we have M row observations, N column attributes. Attributes[k] is a 1 
    dimensional vector, which show all observation values of kth attribute.

    Step:
        1. calculate mean of every attributes' observation, so it will be N
            mean vector.
        2. use mean value to normalize original observation value.
        3. calculate covariance of normalize attributes, and form a covariance 
            matrix used all covariance that is pairwise calculated by two 
            attributes, and we get a N x N cov matrix.
        4. calculate eigenvalues and eigenvectors of covariance matrix, so we
            got total N eigen value and 20 eigen vectors belongs to each eigen 
            value.
        5. return descending sorted eigen values and vectors 
    """
    mean_vector = np.mean(Attributes, axis=0)
    norm_Attributes = Attributes - mean_vector
    std_data = StandardScaler().fit_transform(Attributes)
    cov_mat = np.cov(std_data, rowvar=False)
    values, vectors = np.linalg.eig(cov_mat)
    values, vectors = (np.array(t) for t in zip(*sorted(zip(values, vectors), reverse=True)))
    '''for test
    print('Attribute shape: ', Attributes.shape)
    print('mean_vector shape: ', mean_vector.shape)
    print('Attribute transpose shape: ', Attributes.T.shape)
    print('norm Attribute shape: ', norm_Attributes.shape)
    print('covariance matrix shape: ', cov_mat.shape)
    print('eigen values shape: ', values.shape)
    print('eigen vectors shape: ', vectors.shape)
    #'''
    return mean_vector, values, vectors, norm_Attributes



def scree_graph(values):
    """
    we use pyplot to draw a scree graph, to see how how many components we 
    should use.
    Need: from matplotlib import pyplot
    Args:
        values: eigen values, which is calculated from a covariance matrix.
    Step:
        1. compute values have n length, and create a n+1 array.
        2. set proper tick of x and y axis and label them
        3. draw a 2 dimensional dot-line chart
    """
    sing_vals = np.arange(values.shape[0]) + 1
    pyplot.xticks(np.arange(0, 21, 1))
    #pyplot.yticks(np.arange(-1, 1, 0.05))
    pyplot.xlabel('Principal Component')
    pyplot.ylabel('Eigenvalues')
    pyplot.plot(sing_vals, values, 'ro-', linewidth=2)
    pyplot.show()

def choose_k_largest_eigen_transform(Attributes, vectors, mean_vector, k):

    W = vectors[:, :k]

    
    z = np.dot(Attributes, W)
    print('the results of dimension reduction')
    print(z)

    arr = np.array(z)
    n1 = arr.shape[0]
    xcord1=[];ycord1=[]
    for i in range(n1):
        xcord1.append(arr[i,0])
        ycord1.append(arr[i,1])
    fig = pyplot.figure()
    pyplot.xlabel('First Eigenvetor')
    pyplot.ylabel('Second Eigenvetor')
    pyplot.scatter(xcord1[:78], ycord1[:78], c='red', marker='o')
    pyplot.scatter(xcord1[78:], ycord1[78:], c='blue', marker='o')
    pyplot.savefig('pca.png',dpi=400)
    pyplot.show()
    
    return 


def loadData(file):
    wb = load_workbook(file)
    ws = wb.active

    Features = np.zeros(110)
    F_Name = []
    for k in range(2,22):
        f = np.array([row[k].value for i, row 
            in enumerate(ws.iter_rows()) if i > 1], dtype='float32')
        name = [row[k].value for i, row 
            in enumerate(ws.iter_rows()) if i == 1]

        Features = np.vstack((Features, f))
        F_Name.append(name[0][1:-1])
    Features = np.delete(Features, np.s_[0:1], axis=0)

    return Features.T

Features = loadData(args.XLSX)

mean_vector, e_values, e_vectors, norm_Features = calculate_eigen_of_cov_mat(Features)
scree_graph(e_values)
k = 2
choose_k_largest_eigen_transform(norm_Features, e_vectors, mean_vector, k)


